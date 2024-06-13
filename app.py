import openpyxl
from openpyxl.styles import Font

adress = 'D:\\2. рабочие папки\\AFL\\Аттачи\\31.05.24\\Лист_20.xlsx'

wb = openpyxl.load_workbook(adress)

# print(wb.sheetnames)
for sheet_name in wb.sheetnames: # цикл для итерации по всем листам (sheets) в открытом Excel файле
    # print(sheet_name)

    sheet = wb[sheet_name] # Получаю доступ к конкретному листу по его имени

    res = 0
    i = 1
    for row in sheet.iter_rows():
        if i <= 6:
            i += 1
            continue
        else:
            for cell in row: # if not cell.value and...
                if cell.value is not None and len(cell.value) > 0 and cell.value[0].isalpha() and cell.column_letter not in ["A", "M", "O"]:
                    cell.value = cell.value[0].upper() + cell.value[1:]
                    cell.font = Font(name='Calibri', size=8)
                res += 1
    print(f"На странице {sheet_name} исправлено {res} строк")

wb.save(adress)

# xml - скрипты - глянуть для форматирования xl (libro office)

# Открываем файл Excel
# Проходим по всем листам файла
# Проходим по всем ячейкам на листе
# Проверяем, что ячейка не пустая и начинается с буквы
# Заменяем первую букву на заглавную
# Сохраняем изменения в файле Excel

# try:
                #     if cell.data_type == 'd':
                #         continue
                # except:
                #     pass